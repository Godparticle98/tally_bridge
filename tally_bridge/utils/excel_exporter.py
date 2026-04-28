"""
tally_bridge/utils/excel_exporter.py
Exports ERPNext accounting data to Excel format for manual Tally import.
"""

import frappe
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from frappe.utils import flt, cstr
import io
import os


def _header_style(ws, row_num, headers, fill_color="1F4E79"):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=fill_color)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row_num].height = 30


def _data_row(ws, row_num, values):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_excel_export(from_date=None, to_date=None, company=None):
    """Generate a multi-sheet Excel workbook with all accounting data."""
    company = company or frappe.defaults.get_user_default("Company")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sales Invoices ──────────────────────────────────────────────
    ws_si = wb.create_sheet("Sales Invoices")
    si_headers = ["Invoice No", "Date", "Customer", "Net Total", "Tax Amount",
                  "Grand Total", "Currency", "Status", "Remarks"]
    _header_style(ws_si, 1, si_headers)

    si_filters = {"docstatus": 1, "company": company}
    if from_date and to_date:
        si_filters["posting_date"] = ["between", [from_date, to_date]]

    sales_invoices = frappe.get_all(
        "Sales Invoice", filters=si_filters,
        fields=["name", "posting_date", "customer_name", "net_total",
                "total_taxes_and_charges", "grand_total", "currency", "status", "remarks"]
    )
    for r, inv in enumerate(sales_invoices, 2):
        _data_row(ws_si, r, [
            inv.name, str(inv.posting_date), inv.customer_name,
            flt(inv.net_total, 2), flt(inv.total_taxes_and_charges, 2),
            flt(inv.grand_total, 2), inv.currency, inv.status,
            cstr(inv.remarks)
        ])
    _auto_width(ws_si)

    # ── Purchase Invoices ────────────────────────────────────────────
    ws_pi = wb.create_sheet("Purchase Invoices")
    pi_headers = ["Invoice No", "Supplier Bill No", "Date", "Supplier",
                  "Net Total", "Tax Amount", "Grand Total", "Currency", "Status"]
    _header_style(ws_pi, 1, pi_headers, fill_color="4472C4")

    pi_filters = {"docstatus": 1, "company": company}
    if from_date and to_date:
        pi_filters["posting_date"] = ["between", [from_date, to_date]]

    purchase_invoices = frappe.get_all(
        "Purchase Invoice", filters=pi_filters,
        fields=["name", "bill_no", "posting_date", "supplier_name", "net_total",
                "total_taxes_and_charges", "grand_total", "currency", "status"]
    )
    for r, inv in enumerate(purchase_invoices, 2):
        _data_row(ws_pi, r, [
            inv.name, cstr(inv.bill_no), str(inv.posting_date), inv.supplier_name,
            flt(inv.net_total, 2), flt(inv.total_taxes_and_charges, 2),
            flt(inv.grand_total, 2), inv.currency, inv.status
        ])
    _auto_width(ws_pi)

    # ── Payment Entries ──────────────────────────────────────────────
    ws_pe = wb.create_sheet("Payment Entries")
    pe_headers = ["Payment No", "Type", "Date", "Party Type", "Party",
                  "Paid Amount", "Received Amount", "Mode", "Reference No"]
    _header_style(ws_pe, 1, pe_headers, fill_color="375623")

    pe_filters = {"docstatus": 1, "company": company}
    if from_date and to_date:
        pe_filters["posting_date"] = ["between", [from_date, to_date]]

    payments = frappe.get_all(
        "Payment Entry", filters=pe_filters,
        fields=["name", "payment_type", "posting_date", "party_type",
                "party_name", "paid_amount", "received_amount",
                "mode_of_payment", "reference_no"]
    )
    for r, pay in enumerate(payments, 2):
        _data_row(ws_pe, r, [
            pay.name, pay.payment_type, str(pay.posting_date),
            pay.party_type, pay.party_name,
            flt(pay.paid_amount, 2), flt(pay.received_amount, 2),
            cstr(pay.mode_of_payment), cstr(pay.reference_no)
        ])
    _auto_width(ws_pe)

    # ── Journal Entries ──────────────────────────────────────────────
    ws_je = wb.create_sheet("Journal Entries")
    je_headers = ["JV No", "Date", "Account", "Debit", "Credit",
                  "Party", "Remarks"]
    _header_style(ws_je, 1, je_headers, fill_color="7030A0")

    je_filters = {"docstatus": 1, "company": company}
    if from_date and to_date:
        je_filters["posting_date"] = ["between", [from_date, to_date]]

    journals = frappe.get_all(
        "Journal Entry", filters=je_filters,
        fields=["name", "posting_date", "user_remark"]
    )
    row_num = 2
    for jv in journals:
        doc = frappe.get_doc("Journal Entry", jv.name)
        for acc in doc.accounts:
            _data_row(ws_je, row_num, [
                jv.name, str(jv.posting_date), acc.account,
                flt(acc.debit_in_account_currency, 2),
                flt(acc.credit_in_account_currency, 2),
                cstr(acc.party), cstr(jv.user_remark)
            ])
            row_num += 1
    _auto_width(ws_je)

    # ── Chart of Accounts ────────────────────────────────────────────
    ws_coa = wb.create_sheet("Chart of Accounts")
    coa_headers = ["Account Name", "Account Number", "Type", "Root Type",
                   "Parent Account", "Is Group", "Currency"]
    _header_style(ws_coa, 1, coa_headers, fill_color="833C00")

    accounts = frappe.get_all(
        "Account", filters={"company": company},
        fields=["account_name", "account_number", "account_type",
                "root_type", "parent_account", "is_group", "account_currency"]
    )
    for r, acc in enumerate(accounts, 2):
        _data_row(ws_coa, r, [
            acc.account_name, cstr(acc.account_number), cstr(acc.account_type),
            acc.root_type, cstr(acc.parent_account),
            "Yes" if acc.is_group else "No", cstr(acc.account_currency)
        ])
    _auto_width(ws_coa)

    # ── Save to bytes ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
